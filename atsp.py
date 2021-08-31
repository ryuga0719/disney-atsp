# -*- coding: utf-8 -*-
"""
Created on Thu Feb 18 22:32:30 2021

@author: ryuga
"""

import random
import math
import gurobipy as gu
import networkx as nx
import matplotlib.pyplot as plt
import sys
import openpyxl as px

# ユークリッド距離を計算する関数
def calculate_distance(node1, node2):
    return round(math.sqrt(((node1[0] - node2[0])**2 + (node1[1] - node2[1])**2)), 1)


def calculate_coordinateY(x):
    return round( math.sqrt(500**2 - x**2) , 1)

# excelを読み込む
wb = px.load_workbook("data.xlsx", data_only=True)
# ワークシートを読み込む
ws = wb.worksheets[0]
# ワークシートの最終行
max_row = ws.max_row


# 顧客数
n = max_row - 1


# 座標群
coordinate = {}
# 入場ゲート
for i in range(n):
    coordinate[i] = (ws.cell(i+2,5).value, ws.cell(i+2,6).value)


# 歩く速度(m/m)
walkSpeed = 60

# 距離の辞書
c = {}
for i in range(n):
    for j in range(n):
        c[i,j] = calculate_distance(coordinate[i],coordinate[j])

# 時間の辞書
t = {}
for i in range(n):
    for j in range(n):
        t[i,j] = round(c[i,j]/walkSpeed,1)

# アトラクションを楽しめる時間(サービスタイム)
serviceTime = []
for i in range(n):
    serviceTime.append(ws.cell(i+2,7).value)
  
    
# アトラクションの待ち時間
waitingTime = []
for i in range(n):
    waitingTime.append(ws.cell(i+2,8).value)
    
# 乗り物の乗りたさ
weight = []
for i in range(n):
    weight.append(ws.cell(i+2,9).value)


# 最適化関数
def mtz_strong(n, c, t, walkSpeed, serviceTime, waitingTime, weight):
    model = gu.Model("tsp")
    x,y,z = {},{},{}
    for i in range(n):
        for j in range(n):
            if i != j:
                x[i,j] = model.addVar(vtype="B", name=f"枝:x[{i},{j}]") # 枝ijを使用するか否かを表す0-1変数
                y[i,j] = model.addVar(vtype="C", name=f"時刻：y[{i},{j}]") #jの到着時刻
                z[i,j] = model.addVar(vtype="C", name=f"移動距離:z[{i},{j}]") #jまでの移動距離
                
    model.update()
    
    # -----次数制約 -----
    # departには1本の枝が入出する
    model.addConstr(gu.quicksum(x[0,j] for j in range(1,n)) == 1)
    model.addConstr(gu.quicksum(x[j,0] for j in range(1,n)) == 1)
        
    #departを除いた各ノードには2つの枝が接続される制約
    for i in range(1,n):
        model.addConstr(gu.quicksum(x[i,j] - x[j,i]\
                                        for j in range(n)\
                                        if j != i) == 0)
        model.addConstr(gu.quicksum(x[i,j] for j in range(n) if j != i) <= 1)

          
    # -----到着時間に関する制約 -----
    # ノードjの発時刻の方が、ノードiの発時刻よりも遅いことを表す制約
    for j in range(1, n):
        model.addConstr(gu.quicksum(y[i,j] for i in range(n) if i != j) \
                     + gu.quicksum( (t[j,k] + serviceTime[k] + waitingTime[k]) * x[j,k] - y[j,k]\
                                   for k in range(n) if k != j)
                     <= 0, name=f'y_{j}')
    
    # 出発点と接続されているノードに関する制約
    model.addConstr(gu.quicksum((t[0,k] + serviceTime[k] + waitingTime[k]) * x[0, k] - y[0,k] \
                                          for k in range(1, n))
                            <= 0, name=f'y_{j}')

    # x[i,j]が1の時だけ、y[i,j]が正の値がとれること
    for i in range(n):
         for j in range(n):
             if i != j:
                 model.addConstr(y[i,j] <= 10000 * x[i,j])
                
    # -----移動距離に関する制約 -----
    # ノードjまでの移動距離の方が、ノードiまでの移動距離よりも遅いことを表す制約
    for j in range(1, n):
        model.addConstr(gu.quicksum(z[i,j] for i in range(n) if i != j) \
                     + gu.quicksum( c[j,k] * x[j,k] - z[j,k]\
                                   for k in range(n) if k != j)
                     <= 0, name=f'y_{j}')
    
    # 出発点と接続されているノードに関する制約
    model.addConstr(gu.quicksum(c[0,k] * x[0, k] - z[0,k] \
                                          for k in range(1, n))
                            <= 0, name=f'y_{j}')

    # x[i,j]が1の時だけ、z[i,j]が正の値がとれること
    for i in range(n):
         for j in range(n):
             if i != j:
                 model.addConstr(z[i,j] <= 10000 * x[i,j]) 
    
    
    # 遊ぶ時間の最大を制限する制約
    model.addConstr(gu.quicksum((t[i,j]  + serviceTime[j] + waitingTime[j]) * x[i,j] for (i,j) in x) <= 660)
    
    
        
    #目的関数
    model.setObjective(gu.quicksum(weight[j] * (t[i,j] + serviceTime[j] + waitingTime[j]) * x[i,j]
    for (i,j) in x)
    -gu.quicksum(y[i,0] for i in range(1,n))
    -0.5 * gu.quicksum(z[i,0] for i in range(1,n))
    , gu.GRB.MAXIMIZE)
    model.update()
    model.__data = x,y,z
    return model

    

# solve           
if __name__ == "__main__":
    model = mtz_strong(n, c, t, walkSpeed, serviceTime, waitingTime, weight)
    model.optimize()
    x = model.__data[0]
    y = model.__data[1]
    z = model.__data[2]

    # pathのリスト
    nodesPair = [(i,j) for (i,j) in x if x[i,j].X > 0.5]
    connectedNodes = [[] for i in range(n)]
    for (i,j) in nodesPair:
        connectedNodes[i].append(j)
        connectedNodes[j].append(i)
    
    path = []
    path.append(0)
    i = 0
    j = 0 # 一つ前の点
    while True:
        # connectedNodes[i][0]が前のノードでなかったら
        if connectedNodes[i][0] != j:
            j = i
            i = connectedNodes[i][0]
            # connectedNodes[i][0]が前の点jだったら
            # connectedNodes[i][0]はすでに通過したノードを記しているため
            # connectedNodes[i][1]が次に通過するノードである
        else:
            j = i
            i = connectedNodes[i][1]
                 
        if i == 0:
            path.append(0)
        else:
            path.append(i)
                 
        if i == 0:
            connectedNodes[0].pop(0)
            break
      
    
    plotG = nx.Graph()
    # ノードの追加
    for i in range(n):
        plotG.add_node(i)
    # エッジの追加
    for i in range(len(path)-1):
        plotG.add_edge(path[i],path[i + 1])
    
    plt.figure(figsize=(15,11.5))
    #nx.draw_networkx(plotG,pos=pos,node_color="k",edge_color="k",font_color="w")
    nx.draw_networkx(plotG,pos=coordinate,edge_color="r",font_color="w",width=5)
    plt.show()
    
    # アトラクション順路
    attrName = []
    for i in path:
        attrName.append(ws.cell(i+2,2).value)
        
    route = ""
    for num,i in enumerate(attrName):
        if num == len(attrName)-1:
            route = route + str(i)
            break
        else:
            route = route + str(i) + " -> "
        
    
        
    
    print("######目的関数値######")
    print(model.ObjVal)
    print("######経路は######")
    print(path)
    print(route)
    print("######最終到着時間######")
    arrivalTime = 0
    
    for (i,j) in y:
        if i != 0:
            if y[i,0].X > 0.5:
                arrivalTime = y[i,0].X
                break
    
    print(arrivalTime)
    print("######最終到着時間######")
    move = 0
    for (i,j) in z:
        if i!= 0:
            if z[i,0].X > 0.5:
                move = z[i,0].X
                break
    print(move)
    